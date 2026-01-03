"""
ExcelParser: XLSX 模板解析与填充

解析 Excel 文档中的 {{KEYWORD}} 占位符，并在保留格式的情况下替换内容。
"""

import re
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from typing import Dict, List, Optional
import copy


class ExcelParser:
    """XLSX 模板解析器"""
    
    PLACEHOLDER_PATTERN = re.compile(r'\{\{(\w+)\}\}')
    
    def __init__(self, template_path: str):
        """
        初始化 Excel 解析器
        
        Args:
            template_path: Excel 模板文件路径
        """
        self.template_path = template_path
        self.workbook = load_workbook(template_path)
    
    def find_placeholders(self) -> List[str]:
        """
        查找模板中所有的占位符
        
        Returns:
            占位符名称列表（不包含双花括号）
        """
        placeholders = set()
        
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        matches = self.PLACEHOLDER_PATTERN.findall(cell.value)
                        placeholders.update(matches)
        
        return list(placeholders)
    
    def fill_placeholders(self, content_map: Dict[str, str]) -> None:
        """
        填充占位符，保留原有格式
        
        Args:
            content_map: 占位符名称 -> 填充内容 的映射
        """
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        new_value = cell.value
                        for placeholder, content in content_map.items():
                            pattern = f'{{{{{placeholder}}}}}'
                            if pattern in new_value:
                                new_value = new_value.replace(pattern, content)
                        if new_value != cell.value:
                            # 保留单元格的原有格式
                            cell.value = new_value
    
    def save(self, output_path: str) -> None:
        """
        保存填充后的文档
        
        Args:
            output_path: 输出文件路径
        """
        self.workbook.save(output_path)
    
    def close(self) -> None:
        """关闭工作簿"""
        self.workbook.close()


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        parser = ExcelParser(sys.argv[1])
        placeholders = parser.find_placeholders()
        print(f"Found placeholders: {placeholders}")
        parser.close()
